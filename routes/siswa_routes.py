from flask import Blueprint, session, render_template, request, redirect, url_for, flash, jsonify
from utils.db import db
from utils.decorators import roles_required
import sqlite3
from datetime import datetime
import pandas as pd
import io
from flask import send_file

siswa_bp = Blueprint("siswa", __name__)

def normalize_date(value):
    """
    Terima berbagai format tanggal:
    - YYYY-MM-DD
    - DD-MM-YYYY
    - MM/DD/YYYY
    - DD/MM/YYYY
    - Excel serial date
    Return: YYYY-MM-DD atau None
    """
    if value is None or str(value).strip() == "":
        return None

    # jika sudah datetime (Excel sering begini)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    value = str(value).strip()

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d"
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    raise ValueError(f"Format tanggal tidak valid: {value}")

def format_tanggal_indonesia(tanggal_str):
    if not tanggal_str:
        return "-"
    try:
        tgl = datetime.strptime(tanggal_str, "%Y-%m-%d")
        return tgl.strftime("%d-%m-%Y")
    except Exception:
        return tanggal_str

# ==============================
# HALAMAN UTAMA SISWA
# ==============================
@siswa_bp.route("/students")
def index():

    with db() as d:
        siswa_list = d.execute("""
            SELECT
                s.id,
                s.nis,
                s.nisn,
                s.nama,
                s.tempat_lahir,
                s.tanggal_lahir,
                s.tingkat_default AS tingkat,
                s.nama_ayah,
                s.nama_ibu,
                s.status_masuk,
                s.status,
                k.sub_kelas
            FROM siswa s
            LEFT JOIN kelas_siswa ks ON ks.siswa_id = s.id
            LEFT JOIN kelas k ON k.id = ks.kelas_id
            WHERE s.status = 'aktif'
            ORDER BY s.nama ASC
        """).fetchall()
    siswa_list = [dict(s) for s in siswa_list]

    for s in siswa_list:
        tempat = s["tempat_lahir"] or "-"
        tanggal = format_tanggal_indonesia(s["tanggal_lahir"])

        if tanggal:
            s["ttl"] = f"{tempat}, {tanggal}"
        else:
            s["ttl"] = tempat

    return render_template(
        "dashboard.html",
        active_page="siswa",
        siswa_list=siswa_list
    )

@siswa_bp.route("/students/arsip")
def siswa_arsip():

    with db() as d:
        siswa_list = d.execute("""
            SELECT
                s.id,
                s.nis,
                s.nisn,
                s.nama,
                s.tempat_lahir,
                s.tanggal_lahir,
                s.tingkat_default AS tingkat,
                s.nama_ayah,
                s.nama_ibu,
                s.status_masuk,
                s.status,
                k.sub_kelas
            FROM siswa s
            LEFT JOIN kelas_siswa ks ON ks.siswa_id = s.id
            LEFT JOIN kelas k ON k.id = ks.kelas_id
            WHERE s.status != 'aktif'
            ORDER BY s.nama ASC
        """).fetchall()

    siswa_list = [dict(s) for s in siswa_list]

    for s in siswa_list:
        tempat = s["tempat_lahir"] or "-"
        tanggal = format_tanggal_indonesia(s["tanggal_lahir"])

        if tanggal:
            s["ttl"] = f"{tempat}, {tanggal}"
        else:
            s["ttl"] = tempat

    return render_template(
        "dashboard.html",
        active_page="siswa_arsip",
        siswa_list=siswa_list
    )

# ==============================
# HALAMAN TAMBAH SISWA BARU
# ==============================
@siswa_bp.route("/students/add", methods=["GET", "POST"])
def tambah_siswa():

    if request.method == "POST":

        nis = request.form.get("nis")
        nisn = request.form.get("nisn")
        nik = request.form.get("nik")
        nama = request.form.get("nama")
        jenis_kelamin = request.form.get("jenis_kelamin")
        tempat_lahir = request.form.get("tempat_lahir")
        tanggal_lahir = normalize_date(request.form.get("tanggal_lahir"))

        tingkat = request.form.get("tingkat") or 7
        tahun_masuk = request.form.get("tahun_masuk")
        sekolah_asal = request.form.get("sekolah_asal")

        alamat = request.form.get("alamat")
        desa = request.form.get("desa")
        kecamatan = request.form.get("kecamatan")
        kabupaten = request.form.get("kabupaten")
        provinsi = request.form.get("provinsi")
        nama_ayah = request.form.get("nama_ayah")
        pekerjaan_ayah = request.form.get("pekerjaan_ayah")
        nama_ibu = request.form.get("nama_ibu")
        pekerjaan_ibu = request.form.get("pekerjaan_ibu")
        no_hp = request.form.get("no_hp")

        try:
            with db() as d:
                d.execute("""
                    INSERT INTO siswa 
                    (
                        nis, nisn, nik, nama, jk,
                        tempat_lahir, tanggal_lahir,
                        tingkat_default,
                        tahun_masuk, sekolah_asal,
                        alamat, desa, kecamatan, kabupaten, provinsi,
                        nama_ayah, pekerjaan_ayah,
                        nama_ibu, pekerjaan_ibu,
                        no_hp,
                        status_masuk,
                        status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                        nis,
                        nisn,
                        nik,
                        nama,
                        jenis_kelamin,
                        tempat_lahir,
                        tanggal_lahir,
                        tingkat,
                        tahun_masuk,
                        sekolah_asal,
                        alamat,
                        desa,
                        kecamatan,
                        kabupaten,
                        provinsi,
                        nama_ayah,
                        pekerjaan_ayah,
                        nama_ibu,
                        pekerjaan_ibu,
                        no_hp,
                        "baru",
                        "aktif"
                    ))

            return redirect(url_for("siswa.index"))

        except sqlite3.IntegrityError as e:
            if "siswa.nis" in str(e):
                flash("❌ NIS sudah terdaftar, gunakan NIS lain", "error")
            else:
                flash("❌ Terjadi kesalahan data (duplikat)", "error")

    return render_template(
        "dashboard.html",
        active_page="tambah_siswa",
        siswa=None
    )

@siswa_bp.route("/students/add/pindahan", methods=["GET", "POST"])
def tambah_siswa_pindahan():
    if request.method == "POST":

        nis = request.form.get("nis")
        nisn = request.form.get("nisn")
        nik = request.form.get("nik")
        nama = request.form.get("nama")
        jk = request.form.get("jenis_kelamin")
        tempat_lahir = request.form.get("tempat_lahir")
        tanggal_lahir = normalize_date(request.form.get("tanggal_lahir"))

        asal_sd = request.form.get("asal_sd")
        tahun_lulus_sd = request.form.get("tahun_lulus_sd")

        sekolah_sebelumnya = request.form.get("sekolah_sebelumnya")
        kelas_pindah = request.form.get("kelas_pindah")
        semester_pindah = request.form.get("semester_pindah")
        tanggal_pindah = normalize_date(request.form.get("tanggal_pindah"))

        kelas_diterima = request.form.get("kelas_diterima")
        semester_diterima = request.form.get("semester_diterima")
        tanggal_diterima = normalize_date(request.form.get("tanggal_diterima"))

        alasan_pindah = request.form.get("alasan_pindah")

        alamat = request.form.get("alamat")
        desa = request.form.get("desa")
        kecamatan = request.form.get("kecamatan")
        kabupaten = request.form.get("kabupaten")
        provinsi = request.form.get("provinsi")

        nama_ayah = request.form.get("nama_ayah")
        pekerjaan_ayah = request.form.get("pekerjaan_ayah")
        nama_ibu = request.form.get("nama_ibu")
        pekerjaan_ibu = request.form.get("pekerjaan_ibu")
        no_hp = request.form.get("no_hp")

        try:
            with db() as d:
                d.execute("""
                    INSERT INTO siswa (
                        nis, nisn, nik, nama, jk,
                        tempat_lahir, tanggal_lahir,

                        asal_sd, tahun_lulus_sd,

                        sekolah_sebelumnya,
                        kelas_pindah, semester_pindah,
                        tanggal_pindah,

                        kelas_diterima, semester_diterima,
                        tanggal_diterima,

                        tingkat_default,

                        alamat, desa, kecamatan, kabupaten, provinsi,

                        nama_ayah, pekerjaan_ayah,
                        nama_ibu, pekerjaan_ibu,
                        no_hp,

                        alasan_pindah,
                        status_masuk,
                        status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    nis, nisn, nik, nama, jk,
                    tempat_lahir, tanggal_lahir,

                    asal_sd, tahun_lulus_sd,

                    sekolah_sebelumnya,
                    kelas_pindah, semester_pindah,
                    tanggal_pindah,

                    kelas_diterima, semester_diterima,
                    tanggal_diterima,

                    kelas_pindah or kelas_diterima,

                    alamat, desa, kecamatan, kabupaten, provinsi,

                    nama_ayah, pekerjaan_ayah,
                    nama_ibu, pekerjaan_ibu,
                    no_hp,

                    alasan_pindah,
                    "pindahan",
                    "aktif"
                ))

            flash("Siswa pindahan berhasil ditambahkan", "success")
            return redirect(url_for("siswa.index"))

        except sqlite3.IntegrityError as e:
            flash("❌ Terjadi kesalahan: " + str(e), "error")

    return render_template(
        "dashboard.html",
        active_page="tambah_siswa_pindahan"
    )

# ==============================
# HALAMAN DETAIL SISWA
# ==============================
@siswa_bp.route("/students/<nisn>")
@roles_required("ALL_AUTHENTICATED")
def detail_siswa(nisn):

    with db() as d:
        siswa = d.execute("""
            SELECT
                s.*,

                /* tingkat AKTUAL dari kelas */
                k.tingkat AS tingkat_aktif,
                k.sub_kelas,

                /* fallback jika belum masuk kelas */
                s.tingkat_default AS tingkat_awal

            FROM siswa s
            LEFT JOIN kelas_siswa ks ON ks.siswa_id = s.id
            LEFT JOIN kelas k ON k.id = ks.kelas_id
            WHERE s.nisn = ?
        """, (nisn,)).fetchone()

    if not siswa:
        return "Siswa tidak ditemukan", 404

    siswa = dict(siswa)

    return render_template(
        "dashboard.html",
        active_page="detail_siswa",
        siswa=siswa,
        role=session.get("role")
    )

# ==============================
# DELETE SISWA
# ==============================
@siswa_bp.route("/students/delete/<int:siswa_id>", methods=["POST"])
def delete_siswa(siswa_id):

    with db() as d:

        siswa = d.execute("""
            SELECT status
            FROM siswa
            WHERE id = ?
        """, (siswa_id,)).fetchone()

        if not siswa:
            return {"success": False, "message": "Siswa tidak ditemukan"}

        # cek status
        if siswa["status"] == "aktif":
            return {
                "success": False,
                "message": "Siswa masih aktif, tidak bisa dihapus"
            }

        # cek apakah ada di rombel
        rombel = d.execute("""
            SELECT id
            FROM kelas_siswa
            WHERE siswa_id = ?
        """, (siswa_id,)).fetchone()

        if rombel:
            return {
                "success": False,
                "message": "Siswa masih berada di rombel kelas"
            }

        # jika lolos semua
        d.execute("""
            DELETE FROM siswa
            WHERE id = ?
        """, (siswa_id,))

    return {"success": True}

@siswa_bp.route("/students/ubah-status", methods=["POST"])
def ubah_status():

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Data request tidak valid"
        }), 400

    siswa_id = data.get("id")
    status_baru = data.get("status")

    if not siswa_id or not status_baru:
        return jsonify({
            "success": False,
            "message": "Parameter tidak lengkap"
        }), 400

    with db() as d:

        # cek rombel
        rombel = d.execute("""
            SELECT id
            FROM kelas_siswa
            WHERE siswa_id = ?
        """, (siswa_id,)).fetchone()

        if rombel and status_baru == "nonaktif":
            return jsonify({
                "success": False,
                "message": "Siswa masih berada di rombel. Keluarkan dari rombel terlebih dahulu."
            })

        d.execute("""
            UPDATE siswa
            SET status = ?,
                sekolah_tujuan = NULL,
                alasan_pindah = NULL
            WHERE id = ?
        """, (status_baru, siswa_id))

    return jsonify({
        "success": True,
        "message": "Status siswa berhasil diperbarui"
    })

@siswa_bp.route("/students/pindah", methods=["POST"])
def pindah_siswa():

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Data request tidak valid"
        }), 400

    siswa_id = data.get("id")
    sekolah = data.get("sekolah")
    alasan = data.get("alasan")

    if not siswa_id:
        return jsonify({
            "success": False,
            "message": "ID siswa tidak ditemukan"
        }), 400

    if not sekolah:
        return jsonify({
            "success": False,
            "message": "Sekolah tujuan wajib diisi"
        }), 400

    with db() as d:

        siswa = d.execute("""
            SELECT id, status
            FROM siswa
            WHERE id = ?
        """, (siswa_id,)).fetchone()

        if not siswa:
            return jsonify({
                "success": False,
                "message": "Siswa tidak ditemukan"
            })

        # =========================
        # KELUARKAN DARI ROMBEL
        # =========================
        d.execute("""
            DELETE FROM kelas_siswa
            WHERE siswa_id = ?
        """, (siswa_id,))

        # =========================
        # UPDATE STATUS
        # =========================
        d.execute("""
            UPDATE siswa
            SET status = 'pindah',
                sekolah_tujuan = ?,
                alasan_pindah = ?
            WHERE id = ?
        """, (sekolah, alasan, siswa_id))

    return jsonify({
        "success": True,
        "message": "Siswa berhasil dipindahkan"
    })

@siswa_bp.route("/students/export")
def export_siswa():

    tipe = request.args.get("type")

    if tipe not in ("aktif", "arsip", "semua"):
        tipe = "semua"

    with db() as d:

        base_query = """
            SELECT
                s.nis            AS "NIS",
                s.nisn           AS "NISN",
                s.nama           AS "Nama",
                s.jk             AS "Jenis Kelamin",
                s.tempat_lahir   AS "Tempat Lahir",
                s.tanggal_lahir  AS "Tanggal Lahir",
                s.tingkat_default AS "Tingkat",
                k.sub_kelas      AS "Kelas",
                s.asal_sd         AS "Asal SD",
                s.tahun_lulus_sd  AS "Tahun Lulus",
                s.kelas_diterima  AS "Diterima di Kelas",
                s.semester_diterima AS "Semester Diterima",
                s.tanggal_diterima AS "Tanggal Diterima",
                s.alamat          AS "Alamat Lengkap",
                s.desa            AS "Desa",
                s.kecamatan       AS "Kecamatan",
                s.kabupaten       AS "Kabupaten",
                s.provinsi        AS "Provinsi",
                s.nama_ayah       AS "Nama Ayah",
                s.pekerjaan_ayah  AS "Pekerjaan Ayah",
                s.nama_ibu        AS "Nama Ibu",
                s.pekerjaan_ibu   AS "Pekerjaan Ibu",
                s.no_hp           AS "No HP Orang Tua",
                s.status_masuk   AS "Status Masuk",
                s.status         AS "Status"
            FROM siswa s
            LEFT JOIN kelas_siswa ks ON ks.siswa_id = s.id
            LEFT JOIN kelas k ON k.id = ks.kelas_id
        """

        if tipe == "aktif":
            base_query += " WHERE s.status = 'aktif'"

        elif tipe == "arsip":
            base_query += " WHERE s.status != 'aktif'"

        base_query += " ORDER BY s.nama ASC"

        siswa = d.execute(base_query).fetchall()

    siswa_list = [dict(s) for s in siswa]

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    columns = [
        "NO",
        "NIS",
        "NISN",
        "Nama",
        "Jenis Kelamin",
        "Tempat Lahir",
        "Tanggal Lahir",
        "Tingkat",
        "Kelas",
        "Asal SD",
        "Tahun Lulus",
        "Diterima di Kelas",
        "Semester Diterima",
        "Tanggal Diterima",
        "Alamat Lengkap",
        "Desa",
        "Kecamatan",
        "Kabupaten",
        "Provinsi",
        "Nama Ayah",
        "Pekerjaan Ayah",
        "Nama Ibu",
        "Pekerjaan Ibu",
        "No HP Orang Tua",
        "Status Masuk",
        "Status"
    ]

    if siswa_list:
        df = pd.DataFrame(siswa_list)
        df.insert(0, "NO", range(1, len(df) + 1))
    else:
        # DATA KOSONG → HEADER TETAP ADA
        df = pd.DataFrame(columns=columns)

    # =========================
    # TAMBAH KOLOM NO DI KIRI
    # =========================

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Siswa")
        sheet = writer.book["Data Siswa"]

        # =========================
        # STYLE HEADER
        # =========================
        header_fill = PatternFill("solid", fgColor="FF8C00")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # =========================
        # AUTO WIDTH KOLOM
        # =========================
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[col_letter].width = max_length + 4

    output.seek(0)

    filename = f"data_siswa_{tipe}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==============================
# EDIT DATA SISWA
# ==============================
@siswa_bp.route("/students/<nisn>/edit", methods=["GET", "POST"])
@roles_required("admin", "kepala_sekolah")
def edit_siswa(nisn):

    with db() as d:
        siswa = d.execute("""
            SELECT *
            FROM siswa
            WHERE nisn = ?
        """, (nisn,)).fetchone()

    if not siswa:
        return "Siswa tidak ditemukan", 404

    siswa = dict(siswa)

    if request.method == "POST":
        data = {
            "nis": request.form.get("nis"),
            "nama": request.form.get("nama"),
            "jk": request.form.get("jenis_kelamin"),

            "tempat_lahir": request.form.get("tempat_lahir"),
            "tanggal_lahir": normalize_date(request.form.get("tanggal_lahir")),

            "kelas_pindah": request.form.get("kelas_pindah"),
            "semester_pindah": request.form.get("semester_pindah"),
            "kelas_diterima": request.form.get("kelas_diterima"),
            "semester_diterima": request.form.get("semester_diterima"),

            "tahun_masuk": request.form.get("tahun_masuk"),
            "asal_sd": request.form.get("sekolah_asal"),

            "alamat": request.form.get("alamat"),
            "desa": request.form.get("desa"),
            "kecamatan": request.form.get("kecamatan"),
            "kabupaten": request.form.get("kabupaten"),
            "provinsi": request.form.get("provinsi"),

            "nama_ayah": request.form.get("nama_ayah"),
            "pekerjaan_ayah": request.form.get("pekerjaan_ayah"),
            "nama_ibu": request.form.get("nama_ibu"),
            "pekerjaan_ibu": request.form.get("pekerjaan_ibu"),
            "no_hp": request.form.get("no_hp"),
        }

        try:
            with db() as d:
                d.execute("""
                    UPDATE siswa
                    SET
                        nis = ?,
                        nama = ?,
                        jk = ?,

                        tempat_lahir = ?,
                        tanggal_lahir = ?,

                        tahun_masuk = ?,
                        asal_sd = ?,

                        kelas_pindah = ?,
                        semester_pindah = ?,
                        kelas_diterima = ?,
                        semester_diterima = ?,

                        alamat = ?,
                        desa = ?,
                        kecamatan = ?,
                        kabupaten = ?,
                        provinsi = ?,

                        nama_ayah = ?,
                        pekerjaan_ayah = ?,
                        nama_ibu = ?,
                        pekerjaan_ibu = ?,
                        no_hp = ?

                    WHERE id = ?
                """, (
                    data["nis"],
                    data["nama"],
                    data["jk"],

                    data["tempat_lahir"],
                    data["tanggal_lahir"],

                    data["tahun_masuk"],
                    data["asal_sd"],

                    data["kelas_pindah"],
                    data["semester_pindah"],
                    data["kelas_diterima"],
                    data["semester_diterima"],

                    data["alamat"],
                    data["desa"],
                    data["kecamatan"],
                    data["kabupaten"],
                    data["provinsi"],

                    data["nama_ayah"],
                    data["pekerjaan_ayah"],
                    data["nama_ibu"],
                    data["pekerjaan_ibu"],
                    data["no_hp"],

                    siswa["id"]
                ))

            flash("Data siswa berhasil diperbarui", "success")
            return redirect(url_for("siswa.detail_siswa", nisn=nisn))

        except sqlite3.IntegrityError:
            flash("❌ NIS sudah digunakan siswa lain", "error")

    return render_template(
        "dashboard.html",
        active_page="edit_siswa",
        siswa=siswa
    )

@siswa_bp.route("/students/template/<jenis>/<format>")
def download_template(jenis, format):

    if jenis not in ("baru", "pindahan"):
        return "Jenis template tidak valid", 400

    if format not in ("xlsx", "csv"):
        return "Format tidak valid", 400

    # =========================
    # KOLOM SESUAI FORM
    # =========================
    if jenis == "baru":
        columns = {
            "NIS": "nis",
            "NISN": "nisn",
            "NIK": "nik",
            "Nama": "nama",
            "Jenis Kelamin": "jenis_kelamin",
            "Tempat Lahir": "tempat_lahir",
            "Tanggal Lahir": "tanggal_lahir",
            "Tahun Masuk": "tahun_masuk",
            "Sekolah Asal": "sekolah_asal",
            "Alamat Lengkap": "alamat",
            "Desa": "desa",
            "Kecamatan": "kecamatan",
            "Kabupaten": "kabupaten",
            "Provinsi": "provinsi",
            "Nama Ayah": "nama_ayah",
            "Pekerjaan Ayah": "pekerjaan_ayah",
            "Nama Ibu": "nama_ibu",
            "Pekerjaan Ibu": "pekerjaan_ibu",
            "No HP": "no_hp"
        }
        filename_base = "template_import_siswa_baru"

    else:  # pindahan
        columns = {
            "NIS": "nis",
            "NISN": "nisn",
            "NIK": "nik",
            "Nama": "nama",
            "Jenis Kelamin": "jenis_kelamin",
            "Tempat Lahir": "tempat_lahir",
            "Tanggal Lahir": "tanggal_lahir",

            "Asal SD / MI": "asal_sd",
            "Tahun Lulus SD / MI": "tahun_lulus_sd",

            "Sekolah Sebelumnya": "sekolah_sebelumnya",
            "Kelas Pindah": "kelas_pindah",
            "Semester Pindah": "semester_pindah",
            "Tanggal Pindah": "tanggal_pindah",

            "Diterima di Kelas": "kelas_diterima",
            "Semester Diterima": "semester_diterima",
            "Tanggal Diterima": "tanggal_diterima",

            "Alamat Lengkap": "alamat",
            "Desa": "desa",
            "Kecamatan": "kecamatan",
            "Kabupaten": "kabupaten",
            "Provinsi": "provinsi",

            "Nama Ayah": "nama_ayah",
            "Pekerjaan Ayah": "pekerjaan_ayah",
            "Nama Ibu": "nama_ibu",
            "Pekerjaan Ibu": "pekerjaan_ibu",
            "No HP": "no_hp",

            "Alasan Pindah": "alasan_pindah"
        }
        filename_base = "template_import_siswa_pindahan"

    df = pd.DataFrame(columns=columns.keys())

    # =========================
    # EXPORT
    # =========================
    output = io.BytesIO()

    if format == "xlsx":
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Template Import")
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{filename_base}.xlsx"

    else:  # csv
        df.to_csv(output, index=False)
        mimetype = "text/csv"
        filename = f"{filename_base}.csv"

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype
    )

@siswa_bp.route("/students/import/baru", methods=["POST"])
def import_siswa_baru():

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({
            "success": False,
            "message": "File tidak ditemukan"
        }), 400

    # =========================
    # BACA FILE
    # =========================
    try:
        if file.filename.endswith(".xlsx"):
            df = pd.read_excel(file)
        elif file.filename.endswith(".csv"):
            try:
                df = pd.read_csv(file, encoding="utf-8-sig")
            except:
                df = pd.read_csv(file)
        else:
            return jsonify({
                "success": False,
                "message": "Format file harus CSV atau XLSX"
            }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Gagal membaca file: {str(e)}"
        }), 400

    # =========================
    # NORMALISASI HEADER
    # =========================
    COLUMN_MAPPING = {
        "nis": "nis",
        "nisn": "nisn",
        "nik": "nik",
        "nama": "nama",
        "jenis kelamin": "jenis_kelamin",
        "tempat lahir": "tempat_lahir",
        "tanggal lahir": "tanggal_lahir",
        "tahun masuk": "tahun_masuk",
        "sekolah asal": "sekolah_asal",
        "alamat lengkap": "alamat",
        "desa": "desa",
        "kecamatan": "kecamatan",
        "kabupaten": "kabupaten",
        "provinsi": "provinsi",
        "nama ayah": "nama_ayah",
        "pekerjaan ayah": "pekerjaan_ayah",
        "nama ibu": "nama_ibu",
        "pekerjaan ibu": "pekerjaan_ibu",
        "no hp": "no_hp"
    }

    df.columns = (
        df.columns
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .str.lower()
    )
    df = df.rename(columns=COLUMN_MAPPING)

    # =========================
    # VALIDASI KOLOM WAJIB
    # =========================
    required_columns = [
        "nis", "nisn", "nik", "nama",
        "jenis_kelamin", "tempat_lahir",
        "tanggal_lahir", "tahun_masuk"
    ]

    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        return jsonify({
            "success": False,
            "message": f"Kolom wajib tidak lengkap: {', '.join(missing)}"
        }), 400

    sukses = 0
    gagal = []

    with db() as d:
        for i, row in df.iterrows():
            try:
                nik = str(row["nik"]).strip()
                if not nik.isdigit() or len(nik) != 16:
                    raise ValueError("NIK wajib 16 digit")

                nis = str(row["nis"]).strip()
                nisn = str(row["nisn"]).strip()
                if not nis or not nisn:
                    raise ValueError("NIS & NISN wajib diisi")

                d.execute("""
                    INSERT INTO siswa (
                        nis, nisn, nik, nama, jk,
                        tempat_lahir, tanggal_lahir,
                        tingkat_default,
                        tahun_masuk, sekolah_asal,
                        alamat, desa, kecamatan, kabupaten, provinsi,
                        nama_ayah, pekerjaan_ayah,
                        nama_ibu, pekerjaan_ibu,
                        no_hp,
                        status_masuk, status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    nis, nisn, nik, row["nama"], row["jenis_kelamin"],
                    row["tempat_lahir"], normalize_date(row["tanggal_lahir"]),
                    7,  # AUTO KELAS 7
                    row["tahun_masuk"], row.get("sekolah_asal"),
                    row.get("alamat"), row.get("desa"),
                    row.get("kecamatan"), row.get("kabupaten"),
                    row.get("provinsi"),
                    row.get("nama_ayah"), row.get("pekerjaan_ayah"),
                    row.get("nama_ibu"), row.get("pekerjaan_ibu"),
                    row.get("no_hp"),
                    "baru", "aktif"
                ))

                sukses += 1

            except Exception as e:
                gagal.append(f"Baris {i+2}: {str(e)}")

    if sukses and not gagal:
        message = f"{sukses} data siswa berhasil diimport"
    elif sukses and gagal:
        message = f"{sukses} berhasil, {len(gagal)} gagal"
    else:
        message = "Semua data gagal diimport"

    return jsonify({
        "success": sukses > 0,
        "message": message,
        "berhasil": sukses,
        "gagal": gagal
    })

@siswa_bp.route("/students/import/pindahan", methods=["POST"])
def import_siswa_pindahan():

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({
            "success": False,
            "message": "File tidak ditemukan"
        }), 400

    # =========================
    # BACA FILE
    # =========================
    try:
        if file.filename.endswith(".xlsx"):
            df = pd.read_excel(file)
        elif file.filename.endswith(".csv"):
            try:
                df = pd.read_csv(file, encoding="utf-8-sig")
            except:
                df = pd.read_csv(file)
        else:
            return jsonify({
                "success": False,
                "message": "Format file harus CSV"
            }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Gagal membaca file: {str(e)}"
        }), 400

    # =========================
    # MAPPING HEADER
    # =========================
    COLUMN_MAPPING = {
        "nis": "nis",
        "nisn": "nisn",
        "nik": "nik",
        "nama": "nama",
        "jenis kelamin": "jenis_kelamin",
        "tempat lahir": "tempat_lahir",
        "tanggal lahir": "tanggal_lahir",

        "asal sd / mi": "asal_sd",
        "tahun lulus sd / mi": "tahun_lulus_sd",

        "sekolah sebelumnya": "sekolah_sebelumnya",
        "kelas pindah": "kelas_pindah",
        "semester pindah": "semester_pindah",
        "tanggal pindah": "tanggal_pindah",

        "diterima di kelas": "kelas_diterima",
        "semester diterima": "semester_diterima",
        "tanggal diterima": "tanggal_diterima",

        "alamat lengkap": "alamat",
        "desa": "desa",
        "kecamatan": "kecamatan",
        "kabupaten": "kabupaten",
        "provinsi": "provinsi",

        "nama ayah": "nama_ayah",
        "pekerjaan ayah": "pekerjaan_ayah",
        "nama ibu": "nama_ibu",
        "pekerjaan ibu": "pekerjaan_ibu",
        "no hp": "no_hp",

        "alasan pindah": "alasan_pindah"
    }

    df.columns = (
        df.columns
        .str.replace("\ufeff", "", regex=False)  # hapus BOM CSV
        .str.strip()
        .str.lower()
    )
    df = df.rename(columns=COLUMN_MAPPING)

    # =========================
    # KOLOM WAJIB
    # =========================
    required_columns = [
        "nis", "nisn", "nik", "nama",
        "kelas_diterima", "semester_diterima", "tanggal_diterima"
    ]

    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        return jsonify({
            "success": False,
            "message": f"Kolom wajib tidak lengkap: {', '.join(missing)}"
        }), 400

    sukses = 0
    gagal = []

    with db() as d:
        for i, row in df.iterrows():
            try:
                nik = str(row.get("nik")).strip()
                if not nik.isdigit() or len(nik) != 16:
                    raise ValueError("NIK wajib 16 digit")

                nis = str(row.get("nis")).strip()
                nisn = str(row.get("nisn")).strip()
                if not nis or not nisn:
                    raise ValueError("NIS dan NISN wajib diisi")

                tingkat = row.get("kelas_pindah") or row.get("kelas_diterima")

                d.execute("""
                    INSERT INTO siswa (
                        nis, nisn, nik, nama, jk,
                        tempat_lahir, tanggal_lahir,

                        asal_sd, tahun_lulus_sd,
                        sekolah_sebelumnya,
                        kelas_pindah, semester_pindah, tanggal_pindah,
                        kelas_diterima, semester_diterima, tanggal_diterima,

                        tingkat_default,

                        alamat, desa, kecamatan, kabupaten, provinsi,
                        nama_ayah, pekerjaan_ayah,
                        nama_ibu, pekerjaan_ibu,
                        no_hp,

                        alasan_pindah,
                        status_masuk, status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    nis, nisn, nik, row.get("nama"), row.get("jenis_kelamin"),
                    row.get("tempat_lahir"), normalize_date(row.get("tanggal_lahir")),

                    row.get("asal_sd"), row.get("tahun_lulus_sd"),
                    row.get("sekolah_sebelumnya"),
                    row.get("kelas_pindah"), row.get("semester_pindah"), normalize_date(row.get("tanggal_pindah")),
                    row.get("kelas_diterima"), row.get("semester_diterima"), normalize_date(row.get("tanggal_diterima")),

                    tingkat,

                    row.get("alamat"), row.get("desa"), row.get("kecamatan"),
                    row.get("kabupaten"), row.get("provinsi"),
                    row.get("nama_ayah"), row.get("pekerjaan_ayah"),
                    row.get("nama_ibu"), row.get("pekerjaan_ibu"),
                    row.get("no_hp"),

                    row.get("alasan_pindah"),
                    "pindahan", "aktif"
                ))

                sukses += 1

            except Exception as e:
                gagal.append(f"Baris {i+2}: {str(e)}")

    if sukses and not gagal:
        message = f"{sukses} siswa pindahan berhasil diimport"
    elif sukses and gagal:
        message = f"{sukses} berhasil, {len(gagal)} gagal"
    else:
        message = "Semua data siswa pindahan gagal diimport"

    return jsonify({
        "success": sukses > 0,
        "message": message,
        "berhasil": sukses,
        "gagal": gagal
    })

@siswa_bp.route("/students/nonaktifkan", methods=["POST"])
def nonaktifkan_siswa():

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Data request tidak valid"
        }), 400

    siswa_id = data.get("id")
    tanggal_nonaktif = data.get("tanggal_nonaktif")
    alasan = data.get("alasan")

    if not siswa_id or not tanggal_nonaktif:
        return jsonify({
            "success": False,
            "message": "Tanggal non aktif wajib diisi"
        }), 400

    try:
        tanggal_nonaktif = normalize_date(tanggal_nonaktif)
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 400

    with db() as d:

        # cek siswa
        siswa = d.execute("""
            SELECT id, status
            FROM siswa
            WHERE id = ?
        """, (siswa_id,)).fetchone()

        if not siswa:
            return jsonify({
                "success": False,
                "message": "Siswa tidak ditemukan"
            })

        if siswa["status"] != "aktif":
            return jsonify({
                "success": False,
                "message": "Siswa sudah tidak aktif"
            })

        # ❗ WAJIB: keluarkan dari rombel
        d.execute("""
            DELETE FROM kelas_siswa
            WHERE siswa_id = ?
        """, (siswa_id,))

        # update status
        d.execute("""
            UPDATE siswa
            SET
                status = 'nonaktif',
                tanggal_nonaktif = ?,
                alasan_nonaktif = ?
            WHERE id = ?
        """, (
            tanggal_nonaktif,
            alasan,
            siswa_id
        ))

    return jsonify({
        "success": True,
        "message": "Siswa berhasil dinonaktifkan"
    })